#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel导出工具 - 处理数据导出到Excel
"""

import os

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Excel导出器"""

    def __init__(self):
        self.output_file = None
        self.workbook = None
        self.worksheet = None
        self.current_row = 1
        self.headers = []

    def set_output_file(self, file_path, append_mode=False):
        """设置输出文件"""
        self.output_file = file_path

        if append_mode and os.path.exists(file_path):
            # 追加模式，打开现有文件
            try:
                self.workbook = openpyxl.load_workbook(file_path)
                self.worksheet = self.workbook.active

                # 找到最后一行
                self.current_row = self.worksheet.max_row + 1

                # 获取现有表头
                self.headers = []
                for cell in self.worksheet[1]:
                    if cell.value:
                        self.headers.append(cell.value)

            except Exception as e:
                # 如果打开失败，创建新的工作簿
                self._create_new_workbook()
        else:
            # 创建新工作簿
            self._create_new_workbook()

    def _create_new_workbook(self):
        """创建新的工作簿"""
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "提取数据"
        self.current_row = 1
        self.headers = []

    def add_row(self, data_dict):
        """添加一行数据"""
        if not self.workbook:
            raise ValueError("未设置输出文件")

        # 处理表头
        if not self.headers:
            self._write_headers(list(data_dict.keys()))
        else:
            # 检查是否有新的表头
            new_headers = [key for key in data_dict.keys() if key not in self.headers]
            if new_headers:
                # 添加新表头并重写第一行
                self._update_headers(new_headers)

        # 写入数据行
        for col, header in enumerate(self.headers, 1):
            value = data_dict.get(header, "")

            # 处理表格数据（二维列表）
            if isinstance(value, list) and value and isinstance(value[0], list):
                # 表格数据需要特殊处理，创建新的工作表
                sheet_name = header[:31]  # Excel限制工作表名不超过31个字符
                if sheet_name in self.workbook.sheetnames:
                    sheet = self.workbook[sheet_name]
                else:
                    sheet = self.workbook.create_sheet(title=sheet_name)

                # 写入表格数据
                self._write_table_data(sheet, value)

                # 在主工作表的单元格中创建超链接引用该工作表
                cell = self.worksheet.cell(row=self.current_row, column=col)
                cell.value = f"参见工作表: {sheet_name}"
                cell.hyperlink = f"#{sheet_name}!A1"
                cell.style = "Hyperlink"
            else:
                # 普通数据直接写入
                self.worksheet.cell(row=self.current_row, column=col).value = value

        self.current_row += 1

    def _write_headers(self, headers):
        """写入表头"""
        self.headers = headers
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        self.current_row = 2

    def _update_headers(self, new_headers):
        """更新表头"""
        # 保存原有表头，扩展新表头
        original_headers = self.headers.copy()
        self.headers.extend(new_headers)

        # 重写表头行
        for col, header in enumerate(self.headers, 1):
            cell = self.worksheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 填充已有数据的新列为空
        if self.current_row > 2:
            for row in range(2, self.current_row):
                for col, header in enumerate(self.headers, 1):
                    if header not in original_headers:
                        self.worksheet.cell(row=row, column=col).value = ""

    def _write_table_data(self, sheet, table_data):
        """写入表格数据到单独的工作表"""
        # 清空工作表
        for row in sheet.rows:
            for cell in row:
                cell.value = None

        # 写入表格数据
        for row_idx, row_data in enumerate(table_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx).value = cell_value

                # 如果是第一行，设置为粗体
                if row_idx == 1:
                    sheet.cell(row=row_idx, column=col_idx).font = Font(bold=True)

        # 调整列宽
        self._adjust_column_width(sheet)

    def _adjust_column_width(self, sheet):
        """调整工作表的列宽"""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length

            adjusted_width = max_length + 2
            sheet.column_dimensions[column_letter].width = adjusted_width

    def save(self):
        """保存Excel文件"""
        if not self.workbook or not self.output_file:
            raise ValueError("未设置输出文件")

        # 调整主工作表列宽
        self._adjust_column_width(self.worksheet)

        # 保存文件
        try:
            self.workbook.save(self.output_file)
            return True
        except Exception as e:
            raise ValueError(f"保存Excel文件失败: {str(e)}")
